# import openpyxl


# def w_excel(new_data):
#     # Specify the file path of your Excel file
#     file_path = 'op.xlsx'

#     # Load the Excel workbook
#     workbook = openpyxl.load_workbook(file_path)

#     # Choose a specific sheet to work with (by default, it selects the first sheet)
#     sheet = workbook.active

#     # # Data to be inserted as a list of lists
#     # new_data = [
#     #     ['Name', 'Age', 'City'],
#     #     ['John', 30, 'New York'],
#     #     ['Alice', 25, 'Los Angeles'],
#     #     ['Bob', 28, 'Chicago']
#     # ]

#     # Determine the starting cell where you want to insert the data
#     start_row = sheet.max_row + 1  # Insert after the last row in the sheet

#     # Insert the new data into the sheet
#     for row_data in new_data:
#         if len(row_data)>0:
#             sheet.append(row_data)
#         else:
#             continue
#     # Save the changes back to the Excel file
#     workbook.save(file_path)

#     # Close the workbook after writing
#     workbook.close()
#     print("Done")




from openpyxl import Workbook

def sanitize_data(value):
    # Replace non-printable characters and illegal characters
    sanitized_value = ''.join(c if c.isprintable() and c != '\t' and c != '\n' and c != '\r' else ' ' for c in value)
    return sanitized_value

def write_to_excel(data):
    wb = Workbook()
    ws = wb.active

    for row in data:
        sanitized_row = [sanitize_data(value) for value in row]
        ws.append(sanitized_row)

    wb.save("op.xlsx")



# write_to_excel(data)